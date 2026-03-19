# -*- coding: utf-8 -*-
"""
液位检测引擎 - 完整版
提供简洁的检测接口：输入标注数据和帧，输出液位高度数据
"""

import cv2
import csv
import time
import numpy as np
from pathlib import Path

# 导入拆分的模块函数
from .model_detect import (
    validate_device, load_model, cleanup_temp_models, parse_targets
)
from .init_error import (
    calculate_initstatus, process_masks, reset_init_error_state
)

# 导入相机姿态检测模块
try:
    from utils.cameraposition import (
        set_camera_reference, detect_camera_moved, detect_camera_moved_detail, reset_camera_detector
    )
    CAMERA_POSITION_AVAILABLE = True
except ImportError:
    CAMERA_POSITION_AVAILABLE = False


# ==================== 辅助函数 ====================

def get_class_color(class_name):
    """为不同类别分配不同的颜色"""
    color_map = {
        'liquid': (0, 255, 0),
        'foam': (255, 0, 0),
        'air': (0, 0, 255),
    }
    return color_map.get(class_name, (128, 128, 128))


def calculate_foam_boundary_lines(mask):
    """计算foam mask的顶部和底部边界线"""
    if np.sum(mask) == 0:
        return None, None
    y_coords = np.where(mask)[0]
    if len(y_coords) == 0:
        return None, None
    top_y = np.min(y_coords)
    bottom_y = np.max(y_coords)
    return float(top_y), float(bottom_y)


def analyze_multiple_foams(foam_masks, container_pixel_height):
    """分析多个foam，找到可能的液位边界"""
    if len(foam_masks) < 2:
        return None
    foam_boundaries = []
    for i, mask in enumerate(foam_masks):
        top_y, bottom_y = calculate_foam_boundary_lines(mask)
        if top_y is not None and bottom_y is not None:
            foam_boundaries.append({'top_y': top_y, 'bottom_y': bottom_y, 'center_y': (top_y + bottom_y) / 2})
    if len(foam_boundaries) < 2:
        return None
    foam_boundaries.sort(key=lambda x: x['center_y'])
    error_threshold_px = container_pixel_height * 0.1
    for i in range(len(foam_boundaries) - 1):
        if abs(foam_boundaries[i]['bottom_y'] - foam_boundaries[i + 1]['top_y']) <= error_threshold_px:
            return (foam_boundaries[i]['bottom_y'] + foam_boundaries[i + 1]['top_y']) / 2
    return None


def stable_median(data, max_std=1.0):
    """稳健地计算中位数"""
    if len(data) == 0:
        return 0
    data = np.array(data)
    q1, q3 = np.percentile(data, [25, 75])
    iqr = q3 - q1
    data = data[(data >= q1 - 1.5 * iqr) & (data <= q3 + 1.5 * iqr)]
    if len(data) >= 2 and np.std(data) > max_std:
        data = data[np.abs(data - np.median(data)) <= max_std]
    return float(np.median(data)) if len(data) > 0 else 0


# ==================== 卡尔曼滤波输入选择函数 ====================

def _get_kalman_log_enabled():
    """从配置文件读取卡尔曼日志开关状态"""
    try:
        import yaml
        config_path = Path("database/config/default_config.yaml")
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
                return config.get('kalman_result_compilation', 'release') == 'debug'
    except Exception:
        pass
    return False

KALMAN_INPUT_LOG_ENABLED = _get_kalman_log_enabled()  # 从配置文件读取
KALMAN_INPUT_LOG_PATH = "调试结果/调试日志.xlsx"
_kalman_input_log_initialized = False
_kalman_log_data = []  # 缓存日志数据
_kalman_frame_counter = 0  # 帧序号计数器（与调试帧保存使用相同计数）


def _init_kalman_input_log():
    """初始化调试日志（Excel格式）"""
    global _kalman_input_log_initialized, _kalman_log_data, _kalman_frame_counter
    if not KALMAN_INPUT_LOG_ENABLED or _kalman_input_log_initialized:
        return
    _kalman_log_data = []
    _kalman_frame_counter = 0
    _kalman_input_log_initialized = True


def reset_debug_log():
    """重置调试日志（检测开始时调用，与调试帧保存同步重置）"""
    global _kalman_input_log_initialized, _kalman_log_data, _kalman_frame_counter
    # 先保存已有数据
    if _kalman_log_data:
        _save_kalman_log_to_excel()
    # 重置状态
    _kalman_log_data = []
    _kalman_frame_counter = 0
    _kalman_input_log_initialized = False


def _save_kalman_log_to_excel():
    """将缓存的日志数据保存到Excel文件"""
    global _kalman_log_data
    if not _kalman_log_data:
        return
    try:
        import os
        # 确保目录存在
        log_dir = os.path.dirname(KALMAN_INPUT_LOG_PATH)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir, exist_ok=True)
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "调试日志"
        
        # 表头（第一列为帧序号）
        headers = [
            'frame_index', 'raw_observation', 'predicted_height',
            'selected_input', 'kalman_output', 'liquid_line_position', 'input_action',
            'smooth_output', 'is_full'
        ]
        
        # 表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_idx, row_data in enumerate(_kalman_log_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        # 调整列宽 (1.25倍)，新增帧序号列
        column_widths = [15, 20, 20, 17.5, 17.5, 22.5, 30, 17.5, 12.5]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        wb.save(KALMAN_INPUT_LOG_PATH)
    except ImportError:
        # openpyxl未安装，回退到CSV
        import csv
        csv_path = KALMAN_INPUT_LOG_PATH.replace('.xlsx', '.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                'frame_index', 'raw_observation', 'predicted_height',
                'selected_input', 'kalman_output', 'liquid_line_position', 'input_action',
                'smooth_output', 'is_full'
            ])
            writer.writerows(_kalman_log_data)
    except Exception:
        pass


def _log_kalman_process(raw_observation, predicted_height,
                        selected_input, kalman_output, liquid_line_position, input_action,
                        smooth_output, is_full, channel_id=None):
    """写入调试日志（仅记录channel4，Excel格式）"""
    global _kalman_log_data, _kalman_frame_counter
    if not KALMAN_INPUT_LOG_ENABLED:
        return
    # 只记录channel4
    if channel_id != 'channel4':
        return
    try:
        _init_kalman_input_log()
        _kalman_frame_counter += 1
        _kalman_log_data.append([
            _kalman_frame_counter,  # 帧序号
            round(raw_observation, 2), round(predicted_height, 2),
            round(selected_input, 2), round(kalman_output, 2), round(liquid_line_position, 2), input_action,
            round(smooth_output, 2), is_full
        ])
        # 每100条记录保存一次Excel
        if len(_kalman_log_data) % 100 == 0:
            _save_kalman_log_to_excel()
    except Exception:
        pass


def _log_kalman_input(channel_id, area_idx, observation, predicted_height, container_height_mm,
                      error, error_percent, error_threshold,
                      reject_count_before, reject_count_after, selected_input, action_info):
    """写入卡尔曼输入输出选择日志（仅输入选择阶段，向后兼容）"""
    # 此函数保留用于select_kalman_input内部调用，完整日志由_log_kalman_process记录
    pass


def _is_model_anomaly(rejected_buffer, container_height_mm):
    """
    检测是否为模型异常（重复错误检测导致的异常跳变）
    
    判断条件（同时满足）：
    1. 6次被拒绝值的方差 < 容器高度的 5% → 可能是模型重复错误
    2. 6次被拒绝值的中位数 接近容器边界（0 或 满液）→ 可能是边界误判
    
    Args:
        rejected_buffer: 被拒绝的观测值列表（至少6个）
        container_height_mm: 容器高度(mm)
    
    Returns:
        tuple: (is_anomaly, reason)
    """
    if len(rejected_buffer) < 6:
        return False, ""
    
    recent_rejected = rejected_buffer[-6:]
    rejected_array = np.array(recent_rejected)
    
    # 条件1：方差检查（方差 < 容器高度的 5%）
    rejected_std = np.std(rejected_array)
    variance_threshold = container_height_mm * 0.05
    is_low_variance = rejected_std < variance_threshold
    
    # 条件2：边界检查（中位数接近 0 或 满液）
    rejected_median = np.median(rejected_array)
    boundary_threshold = container_height_mm * 0.10  # 10% 范围内视为边界
    is_near_empty = rejected_median < boundary_threshold  # 接近空
    is_near_full = rejected_median > (container_height_mm - boundary_threshold)  # 接近满
    is_near_boundary = is_near_empty or is_near_full
    
    # 同时满足条件1和条件2 → 判定为模型异常
    if is_low_variance and is_near_boundary:
        boundary_type = "空" if is_near_empty else "满"
        reason = f"模型异常(方差{rejected_std:.2f}<{variance_threshold:.2f}, 中位数{rejected_median:.2f}接近{boundary_type})"
        return True, reason
    
    return False, ""


def select_kalman_input(observation, predicted_height, container_height_mm,
                        reject_count, rejected_buffer,
                        error_threshold_percent=30, reject_count_threshold=6,
                        channel_id=None, area_idx=None):
    """
    卡尔曼滤波输入选择函数
    
    根据观测值与预测值的误差，选择卡尔曼滤波的输入值：
    - 误差 <= 30%：使用观测值
    - 误差 > 30%：使用预测值，累计拒绝计数
    - 连续拒绝6次后：检查是否为模型异常，若是则不重置
    
    Args:
        observation: 第i帧观测值
        predicted_height: 第i-1次卡尔曼滤波预测值
        container_height_mm: 容器高度(mm)，用于计算误差百分比
        reject_count: 当前拒绝计数
        rejected_buffer: 被拒绝的观测值缓冲区列表
        error_threshold_percent: 误差阈值百分比，默认30%
        reject_count_threshold: 拒绝次数阈值，默认6次
        channel_id: 通道ID（用于日志）
        area_idx: 区域索引（用于日志）
    
    Returns:
        tuple: (selected_input, new_reject_count, new_rejected_buffer, action_info)
            - selected_input: 选择的输入值
            - new_reject_count: 更新后的拒绝计数
            - new_rejected_buffer: 更新后的缓冲区
            - action_info: 动作描述信息
    """
    reject_count_before = reject_count
    
    # 计算误差百分比
    error = abs(observation - predicted_height)
    error_percent = (error / container_height_mm) * 100 if container_height_mm > 0 else 0
    
    # 误差 <= 30%：使用观测值，清空计数
    if error_percent <= error_threshold_percent:
        selected_input = observation
        new_reject_count = 0
        new_rejected_buffer = []
        action_info = f"接受观测值(误差{error_percent:.1f}%)"
    else:
        # 误差 > 30%：拒绝观测值
        reject_count += 1
        rejected_buffer = rejected_buffer + [observation]
        
        # 检查是否达到6次拒绝
        if reject_count >= reject_count_threshold:
            # 达到6次：检查是否为模型异常
            is_anomaly, anomaly_reason = _is_model_anomaly(rejected_buffer, container_height_mm)
            
            if is_anomaly:
                # 模型异常：不重置，继续使用预测值，清空计数器
                selected_input = predicted_height
                new_reject_count = 0
                new_rejected_buffer = []
                action_info = f"拒绝重置({anomaly_reason})"
            else:
                # 正常跳变：重置为当前观测值
                selected_input = observation
                new_reject_count = 0
                new_rejected_buffer = []
                action_info = f"强制重置(第{reject_count}次拒绝后重置)"
        else:
            # 未达到6次：使用预测值
            selected_input = predicted_height
            new_reject_count = reject_count
            new_rejected_buffer = rejected_buffer
            action_info = f"拒绝观测值({reject_count}/{reject_count_threshold})"
    
    # 记录日志
    _log_kalman_input(
        channel_id, area_idx, observation, predicted_height, container_height_mm,
        error, error_percent, error_threshold_percent,
        reject_count_before, new_reject_count, selected_input, action_info
    )
    
    return selected_input, new_reject_count, new_rejected_buffer, action_info


# ==================== 液位检测引擎 ====================

class LiquidDetectionEngine:
    """
    液位检测引擎
    
    功能：
        1. 模型加载和YOLO分割推理
        2. 模型推理异常后处理
        3. 分割结果分析 → 液位高度
        4. 多帧时序稳健逻辑
        5. 卡尔曼滤波 + 滑动窗口平滑
        6. 满液状态判断
    
    输入：视频帧 + 标注配置
    输出：液位高度数据字典
    """
    
    def __init__(self, model_path=None, device='cuda', batch_size=4):
        """初始化液位检测引擎"""
        # 模型相关
        self.model = None
        self.model_path = model_path
        self.device = validate_device(device)
        self.batch_size = batch_size
        
        # 标注数据
        self.targets = []
        self.fixed_container_bottoms = []
        self.fixed_container_tops = []
        self.actual_heights = []
        
        # 检测状态
        self.kalman_filters = []
        self.no_liquid_count = []
        self.last_liquid_heights = []
        self.frame_counters = []
        self.consecutive_rejects = []
        self.last_observations = []
        self.last_accepted_observations = []
        self.rejected_observations_buffer = []
        
        # InitError 状态（按通道存储，支持多通道共享引擎）
        self.detect_initstatus = []
        self.init_mask_pixel_counts = []
        self.is_inference_error = []
        
        # 多通道状态存储：{channel_id: {'init_mask_pixel_counts': [], 'is_inference_error': []}}
        self._channel_init_states = {}
        
        # 调试开关
        self.debug = False
        self.modeldebug = True
        
        # 滤波参数
        self.smooth_window = 5
        self.error_percentage = 30  # 硬阈值
        self.soft_error_percentage = 10  # 软阈值
        self.trend_confirm_frames = 3
        self.trend_buffers = []
        self.trend_directions = []
        
        # 满液状态
        self.full_threshold_ratio = 0.9
        self.full_count = []
        self.full_confirm_frames = 3
        
        # 相机姿态检测状态（按通道存储）
        # {channel_id: {'ref_set_time': float, 'camera_status': str, 'last_check_time': float}}
        self._camera_position_states = {}
        self._camera_check_interval = 10.0  # 每10秒重置基准帧检测一次
        # 从配置文件读取相机姿态检测开关
        self._camera_position_enabled = self._get_camera_position_enabled()
    
    def _get_camera_position_enabled(self):
        """从配置文件读取相机姿态检测开关状态"""
        if not CAMERA_POSITION_AVAILABLE:
            return False
        try:
            import yaml
            config_path = Path("database/config/default_config.yaml")
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f)
                    enabled = config.get('camera_position_debug', False)
                    return enabled
        except Exception:
            pass
        return False
    
    def load_model(self, model_path):
        """加载YOLO模型"""
        model, actual_path = load_model(model_path, self.device)
        if model is not None:
            self.model = model
            self.model_path = actual_path
            return True
        return False
    
    def configure(self, boxes, fixed_bottoms, fixed_tops, actual_heights, annotation_initstatus=None):
        """
        配置标注数据
        
        Args:
            boxes: 检测框列表
            fixed_bottoms: 容器底部y坐标列表
            fixed_tops: 容器顶部y坐标列表
            actual_heights: 容器实际高度列表(mm)
            annotation_initstatus: 从标注结果读取的初始状态列表 [0/1/2, ...]
                                   0=默认, 1=满, 2=空
        """
        self.targets = parse_targets(boxes)
        self.fixed_container_bottoms = list(fixed_bottoms)
        self.fixed_container_tops = list(fixed_tops)
        self.actual_heights = [float(h) for h in actual_heights]
        
        num_targets = len(self.targets)
        self._init_state_lists(num_targets)
        self._init_kalman_filters(num_targets)
        
        # 从标注结果读取初始状态（detect_initstatus只从标注结果赋值）
        self.detect_initstatus = calculate_initstatus(num_targets, annotation_initstatus)
        self.init_mask_pixel_counts = [0] * num_targets
        self.is_inference_error = [False] * num_targets
        

    
    def _init_state_lists(self, num_targets):
        """初始化状态列表"""
        self.no_liquid_count = [0] * num_targets
        self.last_liquid_heights = [None] * num_targets
        self.frame_counters = [0] * num_targets
        self.consecutive_rejects = [0] * num_targets
        self.last_observations = [None] * num_targets
        self.last_accepted_observations = [None] * num_targets
        self.rejected_observations_buffer = [[] for _ in range(num_targets)]
        self.trend_buffers = [[] for _ in range(num_targets)]
        self.trend_directions = [None] * num_targets
        self.full_count = [0] * num_targets
    
    def _init_kalman_filters(self, num_targets):
        """初始化一维卡尔曼滤波器（只跟踪位置，响应更快）"""
        self.kalman_filters = []
        for i in range(num_targets):
            kf = cv2.KalmanFilter(1, 1)  # 1维状态，1维观测
            kf.measurementMatrix = np.array([[1]], np.float32)
            kf.transitionMatrix = np.array([[1]], np.float32)
            kf.processNoiseCov = np.array([[0.5]], np.float32)  # 过程噪声
            kf.measurementNoiseCov = np.array([[1]], np.float32)  # 测量噪声
            init_height = self.actual_heights[i] * 0.5 if i < len(self.actual_heights) else 5.0
            kf.statePost = np.array([[init_height]], dtype=np.float32)
            self.kalman_filters.append(kf)
    
    def _get_or_create_channel_init_state(self, channel_id, num_targets, fixed_init_levels=None):
        """获取或创建通道特定的所有状态（包括InitError状态、卡尔曼滤波器等）
        
        Args:
            channel_id: 通道ID
            num_targets: 目标数量
            fixed_init_levels: 初始液位高度列表(mm)，从配置文件读取
            
        Returns:
            dict: 包含所有通道特定状态的字典
        """
        # 标准化 fixed_init_levels
        current_init_levels = list(fixed_init_levels) if fixed_init_levels else []
        
        if channel_id not in self._channel_init_states:
            # 创建通道特定的一维卡尔曼滤波器
            kalman_filters = []
            init_heights_used = []  # 记录实际使用的初始值
            for i in range(num_targets):
                kf = cv2.KalmanFilter(1, 1)  # 1维状态，1维观测
                kf.measurementMatrix = np.array([[1]], np.float32)
                kf.transitionMatrix = np.array([[1]], np.float32)
                kf.processNoiseCov = np.array([[0.5]], np.float32)  # 过程噪声
                kf.measurementNoiseCov = np.array([[1]], np.float32)  # 测量噪声
                # 使用fixed_init_levels作为初始值，如果没有则默认10mm
                init_height = current_init_levels[i] if i < len(current_init_levels) else 10.0
                kf.statePost = np.array([[init_height]], dtype=np.float32)
                kalman_filters.append(kf)
                init_heights_used.append(init_height)
            
            self._channel_init_states[channel_id] = {
                # InitError状态
                'init_mask_pixel_counts': [0] * num_targets,
                'is_inference_error': [False] * num_targets,
                # 卡尔曼滤波状态
                'kalman_filters': kalman_filters,
                'consecutive_rejects': [0] * num_targets,
                'last_observations': [None] * num_targets,
                'last_accepted_observations': [None] * num_targets,
                'rejected_observations_buffer': [[] for _ in range(num_targets)],
                'trend_buffers': [[] for _ in range(num_targets)],
                'trend_directions': [None] * num_targets,
                # 液位检测状态
                'no_liquid_count': [0] * num_targets,
                'last_liquid_heights': [None] * num_targets,
                'frame_counters': [0] * num_targets,
                # 满液状态
                'full_count': [0] * num_targets,
                # 记录配置的初始液位值，用于检测配置变化
                'stored_init_levels': current_init_levels.copy(),
            }
        else:
            # 检测配置变化：比较当前 fixed_init_levels 与已存储的值
            state = self._channel_init_states[channel_id]
            stored_init_levels = state.get('stored_init_levels', [])
            
            if current_init_levels != stored_init_levels:
                # 配置发生变化，更新卡尔曼滤波器初始值并重置相关状态
                kalman_filters = state['kalman_filters']
                for i in range(min(len(kalman_filters), len(current_init_levels))):
                    new_init_height = current_init_levels[i]
                    kalman_filters[i].statePost = np.array([[new_init_height]], dtype=np.float32)
                    kalman_filters[i].statePre = np.array([[new_init_height]], dtype=np.float32)
                
                # 重置卡尔曼相关状态
                state['consecutive_rejects'] = [0] * num_targets
                state['last_observations'] = [None] * num_targets
                state['last_accepted_observations'] = [None] * num_targets
                state['rejected_observations_buffer'] = [[] for _ in range(num_targets)]
                state['trend_buffers'] = [[] for _ in range(num_targets)]
                state['trend_directions'] = [None] * num_targets
                
                # 更新存储的配置值
                state['stored_init_levels'] = current_init_levels.copy()
        
        state = self._channel_init_states[channel_id]
        
        # 确保列表长度足够
        while len(state['init_mask_pixel_counts']) < num_targets:
            state['init_mask_pixel_counts'].append(0)
        while len(state['is_inference_error']) < num_targets:
            state['is_inference_error'].append(False)
        while len(state['kalman_filters']) < num_targets:
            kf = cv2.KalmanFilter(1, 1)  # 1维状态，1维观测
            kf.measurementMatrix = np.array([[1]], np.float32)
            kf.transitionMatrix = np.array([[1]], np.float32)
            kf.processNoiseCov = np.array([[0.5]], np.float32)
            kf.measurementNoiseCov = np.array([[1]], np.float32)
            kf.statePost = np.array([[10.0]], dtype=np.float32)
            state['kalman_filters'].append(kf)
        while len(state['consecutive_rejects']) < num_targets:
            state['consecutive_rejects'].append(0)
        while len(state['last_observations']) < num_targets:
            state['last_observations'].append(None)
        while len(state['last_accepted_observations']) < num_targets:
            state['last_accepted_observations'].append(None)
        while len(state['rejected_observations_buffer']) < num_targets:
            state['rejected_observations_buffer'].append([])
        while len(state['trend_buffers']) < num_targets:
            state['trend_buffers'].append([])
        while len(state['trend_directions']) < num_targets:
            state['trend_directions'].append(None)
        while len(state['no_liquid_count']) < num_targets:
            state['no_liquid_count'].append(0)
        while len(state['last_liquid_heights']) < num_targets:
            state['last_liquid_heights'].append(None)
        while len(state['frame_counters']) < num_targets:
            state['frame_counters'].append(0)
        while len(state['full_count']) < num_targets:
            state['full_count'].append(0)
        
        return state
    
    def _ensure_state_lists_size(self, num_targets):
        """确保状态列表长度足够"""
        while len(self.no_liquid_count) < num_targets:
            self.no_liquid_count.append(0)
        while len(self.last_liquid_heights) < num_targets:
            self.last_liquid_heights.append(None)
        while len(self.frame_counters) < num_targets:
            self.frame_counters.append(0)
        while len(self.consecutive_rejects) < num_targets:
            self.consecutive_rejects.append(0)
        while len(self.last_observations) < num_targets:
            self.last_observations.append(None)
        while len(self.last_accepted_observations) < num_targets:
            self.last_accepted_observations.append(None)
        while len(self.rejected_observations_buffer) < num_targets:
            self.rejected_observations_buffer.append([])
        while len(self.trend_buffers) < num_targets:
            self.trend_buffers.append([])
        while len(self.trend_directions) < num_targets:
            self.trend_directions.append(None)
        while len(self.full_count) < num_targets:
            self.full_count.append(0)
        while len(self.kalman_filters) < num_targets:
            kf = cv2.KalmanFilter(1, 1)  # 1维状态，1维观测
            kf.measurementMatrix = np.array([[1]], np.float32)
            kf.transitionMatrix = np.array([[1]], np.float32)
            kf.processNoiseCov = np.array([[0.5]], np.float32)
            kf.measurementNoiseCov = np.array([[1]], np.float32)
            kf.statePost = np.array([[5.0]], dtype=np.float32)
            self.kalman_filters.append(kf)
    
    def cleanup(self):
        """清理资源"""
        cleanup_temp_models()
        # 保存卡尔曼日志到Excel
        _save_kalman_log_to_excel()

    def analyze_masks_to_height(self, afterinit_mask, container_bottom, 
                                container_pixel_height, container_height_mm, idx,
                                channel_state=None):
        """分析分割结果，计算液位高度
        
        Args:
            channel_state: 通道特定状态字典（用于多通道隔离）
        """
        pixel_per_mm = container_pixel_height / container_height_mm
        liquid_height_mm = None
        
        # 使用通道特定状态或引擎实例状态
        no_liquid_count = channel_state['no_liquid_count'] if channel_state else self.no_liquid_count
        frame_counters = channel_state['frame_counters'] if channel_state else self.frame_counters
        last_liquid_heights = channel_state['last_liquid_heights'] if channel_state else self.last_liquid_heights
        
        liquid_masks = [m for m, c, _ in afterinit_mask if c == 'liquid']
        foam_masks = [m for m, c, _ in afterinit_mask if c == 'foam']
        air_masks = [m for m, c, _ in afterinit_mask if c == 'air']
        
        # 方法1：直接liquid检测
        if liquid_masks:
            topmost_y = float('inf')
            for mask in liquid_masks:
                y_indices = np.where(mask)[0]
                if len(y_indices) > 0:
                    topmost_y = min(topmost_y, np.min(y_indices))
            
            if topmost_y != float('inf'):
                liquid_height_mm = (container_bottom - topmost_y) / pixel_per_mm
                liquid_height_mm = max(0, min(liquid_height_mm, container_height_mm))
                no_liquid_count[idx] = 0
                frame_counters[idx] = 0
                last_liquid_heights[idx] = liquid_height_mm
                return liquid_height_mm
        
        # 未检测到liquid的处理
        no_liquid_count[idx] += 1
        frame_counters[idx] += 1
        
        if no_liquid_count[idx] >= 3:
            # 备选方法2：多foam边界
            if len(foam_masks) >= 2:
                liquid_y = analyze_multiple_foams(foam_masks, container_pixel_height)
                if liquid_y is not None:
                    liquid_height_mm = (container_bottom - liquid_y) / pixel_per_mm
                    liquid_height_mm = max(0, min(liquid_height_mm, container_height_mm))
                    last_liquid_heights[idx] = liquid_height_mm
            
            # 备选方法3：单foam底部
            elif len(foam_masks) == 1:
                _, bottom_y = calculate_foam_boundary_lines(foam_masks[0])
                if bottom_y is not None:
                    liquid_height_mm = (container_bottom - bottom_y) / pixel_per_mm
                    liquid_height_mm = max(0, min(liquid_height_mm, container_height_mm))
                    last_liquid_heights[idx] = liquid_height_mm
            
            # 备选方法4：单air底部
            elif len(air_masks) == 1:
                y_coords = np.where(air_masks[0])[0]
                if len(y_coords) > 0:
                    liquid_height_mm = (container_bottom - np.max(y_coords)) / pixel_per_mm
                    liquid_height_mm = max(0, min(liquid_height_mm, container_height_mm))
                    last_liquid_heights[idx] = liquid_height_mm
            
            # 使用最后液位
            if liquid_height_mm is None and last_liquid_heights[idx] is not None:
                liquid_height_mm = last_liquid_heights[idx]
        else:
            if last_liquid_heights[idx] is not None:
                liquid_height_mm = last_liquid_heights[idx]
        
        if frame_counters[idx] % 3 == 0:
            no_liquid_count[idx] = 0
        
        return liquid_height_mm
    
    def apply_kalman_smooth(self, raw_observation, idx, container_height_mm, channel_state=None, channel_id=None):
        """应用卡尔曼滤波，包含输入选择和校正输出
        
        流程：
        1. 获取卡尔曼预测值
        2. 调用select_kalman_input选择输入（观测值或预测值）
        3. 用选择后的输入进行卡尔曼校正
        4. 滑动窗口平滑输出
        
        Args:
            raw_observation: 原始观测值（来自mask分析的液位高度）
            idx: 目标索引
            container_height_mm: 容器高度(mm)
            channel_state: 通道特定状态字典（用于多通道隔离）
            channel_id: 通道ID（用于调试日志）
        
        Returns:
            tuple: (smooth_height, is_full_confirmed)
        """
        # 使用通道特定状态或引擎实例状态
        kalman_filters = channel_state['kalman_filters'] if channel_state else self.kalman_filters
        full_count = channel_state['full_count'] if channel_state else self.full_count
        consecutive_rejects = channel_state['consecutive_rejects'] if channel_state else self.consecutive_rejects
        rejected_observations_buffer = channel_state['rejected_observations_buffer'] if channel_state else self.rejected_observations_buffer
        
        # 1. 卡尔曼预测
        predicted = kalman_filters[idx].predict()
        predicted_height = float(predicted[0][0])
        
        # 记录输入选择前的拒绝计数
        reject_count_before = consecutive_rejects[idx]
        
        # 2. 输入选择（根据误差决定使用观测值还是预测值）
        # 计算误差用于日志
        error = abs(raw_observation - predicted_height)
        error_percent = (error / container_height_mm) * 100 if container_height_mm > 0 else 0
        
        selected_input, new_reject_count, new_rejected_buffer, action_info = select_kalman_input(
            raw_observation, predicted_height, container_height_mm,
            consecutive_rejects[idx], rejected_observations_buffer[idx],
            error_threshold_percent=self.error_percentage,
            reject_count_threshold=6,
            channel_id=channel_id, area_idx=idx
        )
        
        # 更新拒绝状态
        consecutive_rejects[idx] = new_reject_count
        rejected_observations_buffer[idx] = new_rejected_buffer
        
        # 3. 卡尔曼校正（用选择后的输入）
        kalman_filters[idx].correct(np.array([[selected_input]], dtype=np.float32))
        final_height = float(kalman_filters[idx].statePost[0][0])
        
        # 限制范围
        final_height = max(0, min(final_height, container_height_mm))
        
        # 直接使用卡尔曼输出作为最终结果
        smooth_height = final_height
        
        # 满液判断
        full_threshold_mm = container_height_mm * self.full_threshold_ratio
        is_full = smooth_height >= full_threshold_mm
        
        if is_full:
            full_count[idx] += 1
        else:
            full_count[idx] = 0
        
        is_full_confirmed = full_count[idx] >= self.full_confirm_frames
        
        # 记录完整的卡尔曼输入输出到输出过程日志
        _log_kalman_process(
            raw_observation, predicted_height,
            selected_input, final_height, smooth_height, action_info,
            smooth_height, is_full_confirmed, channel_id
        )
        
        # 记录卡尔曼滤波输出FPS
        try:
            from utils.debug_logger import get_debug_logger
            get_debug_logger().record_kalman_frame(channel_id)
        except:
            pass
        
        return smooth_height, is_full_confirmed
    
    def _check_reset_condition_with_state(self, idx, container_height_mm, rejected_observations_buffer, last_accepted_observations):
        """检查是否应该执行强制重置（使用传入的状态）"""
        rejected_buffer = rejected_observations_buffer[idx]
        last_accepted = last_accepted_observations[idx]
        
        if len(rejected_buffer) < 6 or last_accepted is None:
            return True
        
        rejected_array = np.array(rejected_buffer[-6:])
        rejected_std = np.std(rejected_array)
        rejected_mean = np.mean(rejected_array)
        
        is_stable = rejected_std < container_height_mm * 0.10
        is_large_jump = abs(rejected_mean - last_accepted) > container_height_mm * 0.30
        
        return not (is_stable and is_large_jump)
    
    def _check_trend_confirmation_with_state(self, observation, predicted_height, idx, trend_buffers, trend_directions):
        """检查趋势确认（使用传入的状态）"""
        current_direction = 'up' if observation > predicted_height else 'down'
        last_direction = trend_directions[idx]
        
        if last_direction is None or last_direction != current_direction:
            trend_buffers[idx] = [observation]
            trend_directions[idx] = current_direction
            return 'pending'
        
        trend_buffers[idx].append(observation)
        
        if len(trend_buffers[idx]) >= self.trend_confirm_frames:
            buffer = trend_buffers[idx]
            is_monotonic = True
            for i in range(1, len(buffer)):
                if current_direction == 'up' and buffer[i] < buffer[i-1]:
                    is_monotonic = False
                    break
                if current_direction == 'down' and buffer[i] > buffer[i-1]:
                    is_monotonic = False
                    break
            
            trend_buffers[idx] = []
            trend_directions[idx] = None
            return 'confirmed' if is_monotonic else 'failed'
        
        return 'pending'
    
    def _check_reset_condition(self, idx, container_height_mm):
        """检查是否应该执行强制重置（向后兼容）"""
        return self._check_reset_condition_with_state(
            idx, container_height_mm, 
            self.rejected_observations_buffer, self.last_accepted_observations
        )
    
    def _check_trend_confirmation(self, observation, predicted_height, idx):
        """检查趋势确认（向后兼容）"""
        return self._check_trend_confirmation_with_state(
            observation, predicted_height, idx,
            self.trend_buffers, self.trend_directions
        )
    
    def apply_kalman_predict_only(self, idx, container_height_mm, channel_state=None):
        """检测失败时使用预测值"""
        # 使用通道特定状态或引擎实例状态
        kalman_filters = channel_state['kalman_filters'] if channel_state else self.kalman_filters
        full_count = channel_state['full_count'] if channel_state else self.full_count
        
        predicted = kalman_filters[idx].predict()
        final_height = max(0, min(float(predicted[0][0]), container_height_mm))
        
        # 直接使用卡尔曼预测值
        is_full = final_height >= container_height_mm * self.full_threshold_ratio
        
        if is_full:
            full_count[idx] += 1
        else:
            full_count[idx] = 0
        
        return final_height, full_count[idx] >= self.full_confirm_frames

    # ==================== 相机姿态检测 ====================
    
    def _get_expanded_roi(self, frame, boxes, expand_ratio=0.5):
        """
        计算扩展后的大ROI区域（用于相机姿态检测）
        
        Args:
            frame: 原始帧
            boxes: 标注框列表 [(cx, cy, size), ...]
            expand_ratio: 扩展比例，默认0.5（向外扩展50%）
            
        Returns:
            tuple: (big_roi_image, small_roi_mask, big_roi_coords, small_roi_list)
                - big_roi_image: 大ROI区域图像
                - small_roi_mask: 小ROI区域掩膜（用于排除）
                - big_roi_coords: 大ROI坐标 (x1, y1, x2, y2)
                - small_roi_list: 小ROI坐标列表 [(x1, y1, x2, y2), ...]
        """
        if not boxes or frame is None:
            return None, None, None, []
        
        h, w = frame.shape[:2]
        
        # 计算所有小ROI的边界
        all_x1, all_y1, all_x2, all_y2 = [], [], [], []
        small_roi_list = []  # 保存每个小ROI的坐标
        for cx, cy, size in boxes:
            half = size // 2
            x1 = cx - half
            y1 = cy - half
            x2 = cx + half
            y2 = cy + half
            all_x1.append(x1)
            all_y1.append(y1)
            all_x2.append(x2)
            all_y2.append(y2)
            small_roi_list.append((x1, y1, x2, y2))
        
        # 小ROI的包围盒
        small_x1 = min(all_x1)
        small_y1 = min(all_y1)
        small_x2 = max(all_x2)
        small_y2 = max(all_y2)
        
        # 计算扩展量（向外扩展50%宽度）
        small_width = small_x2 - small_x1
        small_height = small_y2 - small_y1
        expand_x = int(small_width * expand_ratio)
        expand_y = int(small_height * expand_ratio)
        
        # 大ROI坐标（扩展后）
        big_x1 = max(0, small_x1 - expand_x)
        big_y1 = max(0, small_y1 - expand_y)
        big_x2 = min(w, small_x2 + expand_x)
        big_y2 = min(h, small_y2 + expand_y)
        
        # 裁剪大ROI图像
        big_roi_image = frame[big_y1:big_y2, big_x1:big_x2].copy()
        
        # 创建小ROI掩膜（在大ROI坐标系中）
        roi_h, roi_w = big_roi_image.shape[:2]
        small_roi_mask = np.zeros((roi_h, roi_w), dtype=np.uint8)
        
        # 将每个小ROI区域标记为255（排除区域）
        for cx, cy, size in boxes:
            half = size // 2
            # 转换到大ROI坐标系
            sx1 = max(0, (cx - half) - big_x1)
            sy1 = max(0, (cy - half) - big_y1)
            sx2 = min(roi_w, (cx + half) - big_x1)
            sy2 = min(roi_h, (cy + half) - big_y1)
            if sx2 > sx1 and sy2 > sy1:
                small_roi_mask[sy1:sy2, sx1:sx2] = 255
        
        return big_roi_image, small_roi_mask, (big_x1, big_y1, big_x2, big_y2), small_roi_list
    
    def _check_camera_position(self, frame, boxes, channel_id='default'):
        """
        检测相机姿态是否发生变化
        
        Args:
            frame: 当前帧
            boxes: 标注框列表
            channel_id: 通道ID
            
        Returns:
            dict: {
                'status': 'normal' 或 'abnormal',
                'moved': bool,
                'message': str,
                'roi_coords': (x1, y1, x2, y2) 大ROI坐标,
                'inlier_pts': [(x,y), ...] 内点列表（相对于大ROI）,
                'outlier_pts': [(x,y), ...] 外点列表（相对于大ROI）,
                'translation': float 平移量,
                'rotation': float 旋转角度,
                'scale_change': float 尺度变化
            }
        """
        if not self._camera_position_enabled or not CAMERA_POSITION_AVAILABLE:
            return {'status': 'normal', 'moved': False, 'message': '相机检测未启用'}
        
        current_time = time.time()
        
        # 获取或初始化通道状态
        if channel_id not in self._camera_position_states:
            self._camera_position_states[channel_id] = {
                'ref_set_time': 0,
                'camera_status': 'normal',
                'last_check_time': 0,
                'ref_initialized': False
            }
        
        state = self._camera_position_states[channel_id]
        
        # 获取扩展ROI
        big_roi, exclude_mask, roi_coords, small_roi_list = self._get_expanded_roi(frame, boxes)
        if big_roi is None:
            return {'status': 'normal', 'moved': False, 'message': '无法获取ROI'}
        
        # 每10秒重置基准帧
        time_since_ref = current_time - state['ref_set_time']
        if not state['ref_initialized'] or time_since_ref >= self._camera_check_interval:
            # 设置新的基准帧（传入channel_id实现多通道独立检测）
            success = set_camera_reference(big_roi, exclude_mask, channel_id=channel_id)
            if success:
                state['ref_set_time'] = current_time
                state['ref_initialized'] = True
                state['camera_status'] = 'normal'
                if self.debug:
                    print(f"📷 [相机检测] channel={channel_id} 基准帧已重置")
            return {'status': 'normal', 'moved': False, 'message': '基准帧已重置', 'roi_coords': roi_coords, 'small_roi_list': small_roi_list}
        
        # 检测相机是否移动（使用详细版本获取特征点信息）
        detail_result = detect_camera_moved_detail(big_roi, exclude_mask, channel_id=channel_id)
        moved = detail_result['moved']
        
        result = {
            'status': 'abnormal' if moved else 'normal',
            'moved': moved,
            'message': '相机异常' if moved else '相机正常',
            'roi_coords': roi_coords,
            'small_roi_list': small_roi_list,
            'inlier_pts': detail_result.get('inlier_pts', []),
            'outlier_pts': detail_result.get('outlier_pts', []),
            'translation': detail_result.get('translation', 0.0),
            'rotation': detail_result.get('rotation', 0.0),
            'scale_change': detail_result.get('scale_change', 0.0)
        }
        
        if moved:
            state['camera_status'] = 'abnormal'
        else:
            state['camera_status'] = 'normal'
        
        return result

    def detect(self, frame_or_roi_frames, annotation_config=None, channel_id=None):
        """
        检测帧中的液位高度
        
        Args:
            frame_or_roi_frames: 输入数据，支持两种格式：
                - 单个完整帧 (np.ndarray): 内部根据targets裁剪ROI
                - ROI图像列表 (list[np.ndarray]): 预裁剪的ROI图像，跳过内部裁剪
            annotation_config: 可选的标注配置字典
            channel_id: 通道ID（用于多通道状态隔离）
        
        Returns:
            dict: 检测结果 {'liquid_line_positions': {...}, 'success': bool}
        """
        if self.model is None:
            return {'liquid_line_positions': {}, 'success': False}
        
        # 判断输入类型：列表为预裁剪ROI，否则为完整帧
        is_roi_frames = isinstance(frame_or_roi_frames, list)
        
        if annotation_config:
            targets = parse_targets(annotation_config.get('boxes', []))
            fixed_bottoms = annotation_config.get('fixed_bottoms', [])
            fixed_tops = annotation_config.get('fixed_tops', [])
            actual_heights = annotation_config.get('actual_heights', [])
            fixed_init_levels = annotation_config.get('fixed_init_levels', [])  # 读取初始液位
            
            # 从annotation_config中读取init_status，为当前通道创建独立的detect_initstatus
            areas = annotation_config.get('areas', {})
            current_detect_initstatus = []
            if areas:
                for i in range(len(targets)):
                    area_key = f'area_{i+1}'
                    area_info = areas.get(area_key, {})
                    init_status = area_info.get('init_status', 0)
                    current_detect_initstatus.append(init_status)
            else:
                current_detect_initstatus = [0] * len(targets)
            
            # 尝试从annotation_config获取channel_id（如果未传入）
            if channel_id is None:
                channel_id = annotation_config.get('channel_id', 'default')
        else:
            targets = self.targets
            fixed_bottoms = self.fixed_container_bottoms
            fixed_tops = self.fixed_container_tops
            actual_heights = self.actual_heights
            fixed_init_levels = []  # 无配置时为空
            current_detect_initstatus = self.detect_initstatus if self.detect_initstatus else [0] * len(targets)
            if channel_id is None:
                channel_id = 'default'
        
        if not targets:
            return {'liquid_line_positions': {}, 'success': False}
        
        self._ensure_state_lists_size(len(targets))
        
        # 确保current_detect_initstatus长度与targets一致
        while len(current_detect_initstatus) < len(targets):
            current_detect_initstatus.append(0)
        
        # 获取或创建通道特定状态（传入fixed_init_levels用于卡尔曼滤波器初始化）
        channel_init_state = self._get_or_create_channel_init_state(channel_id, len(targets), fixed_init_levels)
        
        try:
            liquid_line_positions = {}
            
            # 根据输入类型处理
            if is_roi_frames:
                # 预裁剪ROI模式：直接使用传入的ROI图像列表
                roi_frames = frame_or_roi_frames
                # 跳过相机姿态检测（需要完整帧）
                camera_status_result = {'status': 'normal', 'moved': False, 'message': ''}
                
                for idx, (center_x, center_y, crop_size) in enumerate(targets):
                    if idx >= len(roi_frames):
                        continue
                    
                    cropped = roi_frames[idx]
                    if cropped is None or cropped.size == 0:
                        continue
                    
                    # 计算原始帧中的坐标（用于结果输出）
                    half_size = crop_size // 2
                    top = center_y - half_size
                    left = center_x - half_size
                    bottom = center_y + half_size
                    right = center_x + half_size
                    
                    result = self._detect_single_target(
                        cropped, idx, top,
                        fixed_bottoms[idx] if idx < len(fixed_bottoms) else None,
                        fixed_tops[idx] if idx < len(fixed_tops) else None,
                        actual_heights[idx] if idx < len(actual_heights) else 20.0,
                        current_detect_initstatus,
                        channel_init_state,
                        channel_id
                    )
                    
                    # 处理结果
                    liquid_line_positions[idx] = self._process_detect_result(
                        result, idx, top, left, right, bottom,
                        fixed_bottoms, fixed_tops, actual_heights,
                        channel_init_state
                    )
            else:
                # 完整帧模式：内部裁剪ROI
                frame = frame_or_roi_frames
                h, w = frame.shape[:2]
                
                # 相机姿态检测（使用boxes信息计算扩展ROI）
                camera_status_result = self._check_camera_position(frame, targets, channel_id)
                
                for idx, (center_x, center_y, crop_size) in enumerate(targets):
                    half_size = crop_size // 2
                    top = max(center_y - half_size, 0)
                    bottom = min(center_y + half_size, h)
                    left = max(center_x - half_size, 0)
                    right = min(center_x + half_size, w)
                    
                    cropped = frame[top:bottom, left:right]
                    if cropped.size == 0:
                        continue
                    
                    result = self._detect_single_target(
                        cropped, idx, top,
                        fixed_bottoms[idx] if idx < len(fixed_bottoms) else None,
                        fixed_tops[idx] if idx < len(fixed_tops) else None,
                        actual_heights[idx] if idx < len(actual_heights) else 20.0,
                        current_detect_initstatus,
                        channel_init_state,
                        channel_id
                    )
                    
                    # 处理结果
                    liquid_line_positions[idx] = self._process_detect_result(
                        result, idx, top, left, right, bottom,
                        fixed_bottoms, fixed_tops, actual_heights,
                        channel_init_state
                    )
            
            detection_result = {
                'liquid_line_positions': liquid_line_positions,
                'success': len(liquid_line_positions) > 0,
                # 相机姿态检测结果
                'camera_status': camera_status_result.get('status', 'normal'),
                'camera_moved': camera_status_result.get('moved', False),
                'camera_message': camera_status_result.get('message', ''),
                # 相机姿态调试绘制数据
                'camera_roi_coords': camera_status_result.get('roi_coords'),
                'camera_small_roi_list': camera_status_result.get('small_roi_list', []),
                'camera_inlier_pts': camera_status_result.get('inlier_pts', []),
                'camera_outlier_pts': camera_status_result.get('outlier_pts', []),
                'camera_translation': camera_status_result.get('translation', 0.0),
                'camera_rotation': camera_status_result.get('rotation', 0.0),
                'camera_scale_change': camera_status_result.get('scale_change', 0.0)
            }
            
            return detection_result
            
        except Exception as e:
            if self.debug:
                print(f"❌ [detect] 异常: {e}")
            return {'liquid_line_positions': {}, 'success': False}
    
    def _process_detect_result(self, result, idx, top, left, right, bottom,
                               fixed_bottoms, fixed_tops, actual_heights, channel_init_state):
        """处理单个目标的检测结果
        
        Args:
            result: _detect_single_target的返回值
            idx: 目标索引
            top, left, right, bottom: ROI坐标
            fixed_bottoms, fixed_tops, actual_heights: 容器参数
            channel_init_state: 通道状态
            
        Returns:
            dict: 格式化的检测结果
        """
        observation_mask = []
        if isinstance(result, tuple):
            if len(result) == 4:
                liquid_height_mm, is_full, error_flag, observation_mask = result
            elif len(result) == 3:
                liquid_height_mm, is_full, error_flag = result
            elif len(result) == 2:
                liquid_height_mm, is_full = result
                error_flag = None
            else:
                liquid_height_mm, is_full, error_flag = result, False, None
        else:
            liquid_height_mm, is_full, error_flag = result, False, None
        
        # 获取容器参数
        container_bottom_y = fixed_bottoms[idx] if idx < len(fixed_bottoms) else 0
        container_top_y = fixed_tops[idx] if idx < len(fixed_tops) else 0
        container_height_mm = actual_heights[idx] if idx < len(actual_heights) else 20.0
        
        if liquid_height_mm is None:
            # 使用卡尔曼预测值作为回退
            kalman_filters = channel_init_state['kalman_filters'] if channel_init_state else self.kalman_filters
            if idx < len(kalman_filters):
                predicted = kalman_filters[idx].predict()
                liquid_height_mm = max(0, min(float(predicted[0][0]), container_height_mm))
            else:
                liquid_height_mm = 0.0
            is_full = False
            error_flag = error_flag or 'detect_zero'
        
        container_pixel_height = container_bottom_y - container_top_y
        pixel_per_mm = container_pixel_height / container_height_mm if container_height_mm > 0 else 1
        height_px = int(liquid_height_mm * pixel_per_mm)
        
        return {
            'y': container_bottom_y - height_px,
            'height_mm': liquid_height_mm,
            'height_px': height_px,
            'left': left, 'right': right, 'top': top, 'bottom': bottom,
            'is_full': is_full,
            'error_flag': error_flag,
            'pixel_per_mm': pixel_per_mm,
            'observation_mask': observation_mask if self.modeldebug else None
        }
    
    def _detect_single_target(self, cropped, idx, crop_top_y, container_bottom, container_top, container_height_mm, current_detect_initstatus=None, channel_init_state=None, channel_id=None):
        """检测单个目标的液位高度
        
        Args:
            cropped: 裁剪后的图像
            idx: 目标索引
            crop_top_y: 裁剪区域顶部y坐标
            container_bottom: 容器底部y坐标
            container_top: 容器顶部y坐标
            container_height_mm: 容器高度(mm)
            current_detect_initstatus: 当前通道的初始状态列表
            channel_init_state: 通道特定的InitError状态字典（包含init_mask_pixel_counts和is_inference_error）
            channel_id: 通道ID（用于日志记录）
        """
        try:
            if container_bottom is None or container_top is None:
                container_bottom = self.fixed_container_bottoms[idx] if idx < len(self.fixed_container_bottoms) else 0
                container_top = self.fixed_container_tops[idx] if idx < len(self.fixed_container_tops) else 0
                container_height_mm = self.actual_heights[idx] if idx < len(self.actual_heights) else 20.0
            
            container_pixel_height = container_bottom - container_top
            
            # YOLO推理
            # TensorRT模型已内置FP16，不需要再设置half参数
            is_tensorrt = self.model_path and self.model_path.endswith('.engine')
            
            # TensorRT推理时不传递device参数，让模型自动使用构建时的设备
            predict_kwargs = {
                'source': cropped,
                'imgsz': 640,
                'conf': 0.5,
                'iou': 0.5,
                'save': False,
                'verbose': False,
                'stream': False
            }
            
            if is_tensorrt:
                # TensorRT模型：不设置half和device，使用模型内置配置
                pass
            else:
                # PyTorch模型：设置device和half
                predict_kwargs['device'] = self.device
                predict_kwargs['half'] = self.device != 'cpu'
            
            results = self.model.predict(**predict_kwargs)
            result = results[0]
            
            if result.masks is None:
                return None, False, 'detect_zero', []
            
            masks = result.masks.data.cpu().numpy() > 0.5
            classes = result.boxes.cls.cpu().numpy().astype(int)
            confidences = result.boxes.conf.cpu().numpy()
            
            afterinit_mask = []
            low_conf_masks = []  # 保存低置信度mask用于绘制
            for i in range(len(masks)):
                resized_mask = cv2.resize(masks[i].astype(np.uint8), 
                                         (cropped.shape[1], cropped.shape[0])) > 0.5
                if confidences[i] >= 0.3:
                    afterinit_mask.append((resized_mask, self.model.names[classes[i]], confidences[i]))
                else:
                    low_conf_masks.append((resized_mask, self.model.names[classes[i]], confidences[i]))
            
            if not afterinit_mask:
                error_flag = 'detect_low' if len(masks) > 0 else 'detect_zero'
                # detect_low时返回低置信度mask用于蓝色绘制，detect_zero时返回空列表
                return None, False, error_flag, low_conf_masks if error_flag == 'detect_low' else []
            
            # InitError处理（使用通道特定的状态，避免多通道共享引擎时状态混乱）
            detect_initstatus_to_use = current_detect_initstatus if current_detect_initstatus else self.detect_initstatus
            
            # 使用通道特定的状态，如果没有则使用引擎实例级别的状态（向后兼容）
            if channel_init_state:
                init_mask_pixel_counts = channel_init_state['init_mask_pixel_counts']
                is_inference_error = channel_init_state['is_inference_error']
            else:
                init_mask_pixel_counts = self.init_mask_pixel_counts
                is_inference_error = self.is_inference_error
            
            # 记录InitError处理前的状态
            init_status_val = detect_initstatus_to_use[idx] if idx < len(detect_initstatus_to_use) else 0
            before_init_count = len(afterinit_mask)
            
            afterinit_mask, init_mask_pixel_counts, is_inference_error = process_masks(
                afterinit_mask, idx, detect_initstatus_to_use,
                init_mask_pixel_counts, is_inference_error,
                channel_id=channel_id
            )
            
            # 更新状态（写回通道特定状态或引擎实例状态）
            if channel_init_state:
                channel_init_state['init_mask_pixel_counts'] = init_mask_pixel_counts
                channel_init_state['is_inference_error'] = is_inference_error
            else:
                self.init_mask_pixel_counts = init_mask_pixel_counts
                self.is_inference_error = is_inference_error
            
            if not afterinit_mask:
                return 0.0, False, 'init_error_corrected', []
            
            # 坐标转换
            container_bottom_in_crop = container_bottom - crop_top_y
            
            # 统计mask类别数量
            liquid_count = len([m for m, c, _ in afterinit_mask if c == 'liquid'])
            foam_count = len([m for m, c, _ in afterinit_mask if c == 'foam'])
            air_count = len([m for m, c, _ in afterinit_mask if c == 'air'])
            
            # 分析mask获取液位（使用通道特定状态）
            raw_height = self.analyze_masks_to_height(
                afterinit_mask, container_bottom_in_crop,
                container_pixel_height, container_height_mm, idx,
                channel_state=channel_init_state
            )
            
            if raw_height is not None:
                # 获取卡尔曼预测值
                kalman_filters = channel_init_state['kalman_filters'] if channel_init_state else self.kalman_filters
                predicted_before = kalman_filters[idx].statePre[0][0] if idx < len(kalman_filters) else 0
                
                height, is_full = self.apply_kalman_smooth(
                    raw_height, idx, container_height_mm,
                    channel_state=channel_init_state,
                    channel_id=channel_id
                )
                
                return height, is_full, None, afterinit_mask
            else:
                kalman_filters = channel_init_state['kalman_filters'] if channel_init_state else self.kalman_filters
                if idx < len(kalman_filters):
                    height, is_full = self.apply_kalman_predict_only(
                        idx, container_height_mm,
                        channel_state=channel_init_state
                    )
                    return height, is_full, 'detect_zero', afterinit_mask
                return None, False, 'detect_zero', afterinit_mask
                
        except Exception as e:
            if self.debug:
                print(f"❌ [目标{idx}] 检测异常: {e}")
            return None, False, 'detect_zero', []
    
    def reset_target(self, target_idx):
        """重置指定目标的状态"""
        if target_idx < len(self.consecutive_rejects):
            self.consecutive_rejects[target_idx] = 0
        if target_idx < len(self.last_observations):
            self.last_observations[target_idx] = None
        if target_idx < len(self.no_liquid_count):
            self.no_liquid_count[target_idx] = 0
        if target_idx < len(self.frame_counters):
            self.frame_counters[target_idx] = 0
